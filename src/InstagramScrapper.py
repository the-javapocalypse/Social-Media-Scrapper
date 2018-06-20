import os, platform, time, urllib.request, openpyxl, operator
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl import Workbook



class InstagramScrapper:
    def Create_Dir(self, dir_name):
        if not os.path.exists("data"):
            try:
                os.mkdir("data")
                print("Created directory 'data'")
            except:
                print("Unable to create directory 'data': Directory already exists")
        else:
            print("Unable to create directory 'data': Directory already exists")

        if not os.path.exists("data/data_" + dir_name):
            try:
                os.mkdir("data/data_" + dir_name)
                print("Created directory 'data/data_" + dir_name + "'")
            except:
                print("Unable to create directory 'data/data_" + dir_name + "': Directory already exists")
        else:
            print("Unable to create directory 'data/data_" + dir_name + "': Directory already exists")

        if not os.path.exists("data/data_" + dir_name + '/img'):
            try:
                os.mkdir("data/data_" + dir_name + '/img')
                print("Created directory 'data/data_" + dir_name + "/img'")
            except:
                print("Unable to create directory 'data/data_" + dir_name + "/img': Directory already exists")
        else:
            print("Unable to create directory 'data/data_" + dir_name + "/img': Directory already exists")

    def Scrape_Instagram(self, tag, limit=20, browser='chrome'):
        self.Create_Dir(tag)

        print("Starting Scrapping Instagram")
        file_path = "data/data_" + tag
        keyword = tag
        # Adding path.
        if not os.getcwd() in os.get_exec_path():
            # print('adding path')
            if platform.system() == "Windows":
                os.environ["PATH"] = os.environ["PATH"] + ";" + os.getcwd()
            else:
                os.environ["PATH"] = os.environ["PATH"] + ":" + os.getcwd()

        # opening instagram in browser
        if 'chrome' in browser.lower():
            driver = webdriver.Chrome()
        else:
            driver = webdriver.Firefox()
        driver.get("https://www.instagram.com/" + "explore/tags/" + tag)

        print("Loading Posts")
        time.sleep(10)
        print("Loading Data")

        # Clicking on load more once to load more images. Afterwards we will just
        # tap space to scroll to the page end to load more images
        actions = ActionChains(driver)
        actions.send_keys(Keys.SPACE).perform()
        actions.send_keys(Keys.SPACE).perform()
        actions.send_keys(Keys.SPACE).perform()

        time.sleep(5)

        # Just tap space to scroll to the page end to load more images
        clear = lambda: os.system('cls')
        msg = "Loading Images"
        class_div_img = ["_si7dy"]
        for div in class_div_img:
            if len(driver.find_elements_by_class_name(div)) > 1:
                while (len(driver.find_elements_by_class_name(div)) ) <= limit :
                    actions.send_keys(Keys.SPACE).perform()
                    msg = msg + "."
                    print(msg)
                    print(len(driver.find_elements_by_class_name(div)))
                    time.sleep(2.5)
                    if len(msg) > 18:
                        msg = "Loading Images"
        print(str(limit) + " images loaded")

        # Storing images links and captions. Slicing with 9 because first 9 posts
        # are the 'most popular' posts for that tag
        img_src = []
        img_caption = []
        hashtags = {}

        for data in driver.find_elements_by_class_name("FFVAD"):
            #     u.get_attribute("href").split("/")[4]
            img_caption.append(data.get_attribute("alt"))
            img_src.append(data.get_attribute("src"))

        img_caption = img_caption[9:limit + 9]
        img_src = img_src[9:limit + 9]
        img_caption.sort()

        # Create a workbook for excel
        tag_File = file_path + "/" + tag + "_Instagram.xlsx"
        wb = openpyxl.Workbook()
        ws_Captions = wb.create_sheet(title="Caption")
        col = 'A'
        row = 1

        print("Dumping data in excel file")
        for caption in img_caption:
            tags = caption.split("#")
            # write caption to excel file
            ws_Captions[col + str(row)] = tags[0]
            row += 1

            # strip tags
            tags = tags[1:]
            for tag in tags:
                cleaned = tag.replace(" ", "").replace("\n", "")
                cleaned = cleaned.lower()
                if cleaned not in hashtags and len(cleaned) < 20:
                    hashtags[cleaned] = 1
                elif cleaned in hashtags and len(cleaned) < 20:
                    hashtags[cleaned] = hashtags[cleaned] + 1

        # sort hashtags with frequencies and store them in excel
        hashtags = sorted(hashtags.items(), key=operator.itemgetter(1), reverse=True)

        ws_Tags = wb.create_sheet(title="Tags")
        tagName = 'A'
        tagFreq = 'B'
        row = 1

        for tag in hashtags:
            ws_Tags[tagName + str(row)] = tag[0]
            ws_Tags[tagFreq + str(row)] = tag[1]
            row += 1

        wb.save(tag_File)

        print("Dumping Images. This will take some time!")
        row = 1
        for src in img_src:
            urllib.request.urlretrieve(src, file_path + '/img/Instagram_' + str(row) + ".jpeg")
            row += 1
            if (row % 10 == 0):
                print("(" + str(row) + "/" + str(len(img_src)) + ") Images Downloaded")

        print("Closing Instagram")
        driver.quit()